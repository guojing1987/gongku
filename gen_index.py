# -*- coding: utf-8 -*-
import sys, os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

SKILL_DIR = '/home/sandbox/.openclaw/workspace/skills/xiaoyi-xlsx'
sys.path.insert(0, os.path.join(SKILL_DIR, 'scripts'))
from cover import PALETTE

P = PALETTE['pure']
thin_border = Border(
    left=Side(style='thin', color=P['border']),
    right=Side(style='thin', color=P['border']),
    top=Side(style='thin', color=P['border']),
    bottom=Side(style='thin', color=P['border']))
header_fill  = PatternFill(start_color=P['header_bg'], end_color=P['header_bg'], fill_type="solid")
header_font  = Font(color="FFFFFF", bold=True, name="Arial", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font    = Font(color=P['text_primary'], name="Arial", size=10)
alt_fill     = PatternFill(start_color=P['bg_alt'], end_color=P['bg_alt'], fill_type="solid")
wraptop      = Alignment(vertical="top", wrap_text=True)

wb = Workbook()

# ============ Sheet1 使用说明 ============
ws0 = wb.active
ws0.title = '使用说明'
ws0.sheet_view.showGridLines = False
ws0['A1'] = '工控元件资料库 · 起步版框架'
ws0['A1'].font = Font(color=P['header_bg'], bold=True, size=16)
ws0['A2'] = '用途：设备管理现场查询——输入故障代码 / 模糊型号，快速定位说明书、故障代码表、维修教程并可下载'
ws0['A2'].font = body_font
rows_desc = [
    ('', ''),
    ('如何使用本表', ''),
    ('1. 现场查故障代码', 'Ctrl+F 输入故障代码（如 E.30 或 SP1），若已收录会精确命中对应表；若仅收录了型号，会引导到对应资料。'),
    ('2. 模糊找型号', '在“索引总表”里筛选“品牌”或“型号(含别名)”列，输入记得的任意字符即可定位（例如输入“汇川”看它全部型号）。'),
    ('3. 下载资料', '“资料位置”列给出官方下载入口或本地文件路径；聊天框问我，我也能直接把对应PDF/表格发到你手机。'),
    ('4. 告诉我常用型号', '你列出的品牌和现场在用型号越多，我收录越快越准。'),
    ('', ''),
    ('品牌官方资料下载入口（免费公开）', ''),
    ('ABB', 'https://www.abb.com.cn  （产品 → 下载中心）'),
    ('西门子', 'https://www.siemens.com/cn  （支持/资料下载）'),
    ('三菱', 'https://www.mitsubishielectric-automation.cn  （下载中心）'),
    ('汇川', 'https://www.inovance.com  （支持与下载）'),
    ('信捷', 'https://www.xinje.com  （下载中心）'),
    ('雷赛', 'https://www.leisai.com  （资料下载）'),
    ('欧姆龙', 'https://www.fa.omron.com.cn  （资料下载）'),
    ('台达', 'https://www.delta-china.com.cn  （资料下载）'),
    ('施耐德', 'https://www.se.com/cn  （资料库/下载）'),
    ('安川', 'https://www.yaskawa.com.cn  （资料下载）'),
    ('', ''),
    ('重要说明', '故障代码含义必须来自原厂手册，绝不能凭经验编造——本库所有“故障代码表/维修教程”列在真实资料入库前统一标注“待收集”，避免误导现场维修。'),
]
r = 4
for a, b in rows_desc:
    ws0.cell(r, 1, a).font = Font(color=P['header_bg'], bold=True) if a and not a.startswith('http') and len(a) > 1 and ' ' not in a[:2] else body_font
    ws0.cell(r, 1, a).font = body_font
    ws0.cell(r, 2, b).font = body_font
    ws0.cell(r, 1).alignment = wraptop
    ws0.cell(r, 2).alignment = wraptop
    r += 1
ws0.column_dimensions['A'].width = 30
ws0.column_dimensions['B'].width = 70

# ============ Sheet2 索引总表 ============
ws = wb.create_sheet('索引总表')
ws.sheet_view.showGridLines = False
headers = ['元件类别', '品牌', '型号(含别名)', '中文说明书', '故障代码表', '维修教程', '资料位置/路径', '收录状态', '备注']
for c, h in enumerate(headers, 1):
    cell = ws.cell(1, c, h)
    cell.fill = header_fill; cell.font = header_font; cell.alignment = header_align
    cell.border = thin_border

# 示例样板行（型号为真实公开型号；资料状态标注待收集，不编造故障代码）
samples = [
    ['PLC', '西门子', 'S7-1200', '待收集', '待收集', '待收集', '西门子官网下载中心', '样板·待收集', '入门级小型PLC，现场常用'],
    ['PLC', '三菱', 'FX3U', '待收集', '待收集', '待收集', '三菱官网下载中心', '样板·待收集', '经典小型机'],
    ['PLC', '汇川', 'H5U', '待收集', '待收集', '待收集', '汇川官网下载中心', '样板·待收集', '国内常见'],
    ['伺服', '安川', 'Σ-V 系列(如SGD7S)', '待收集', '待收集', '待收集', '安川官网下载中心', '样板·待收集', '伺服驱动器常见'],
    ['变频器', 'ABB', 'ACS580', '待收集', '待收集', '待收集', 'ABB官网下载中心', '样板·待收集', '通用变频器'],
    ['变频器', '西门子', 'V20', '待收集', '待收集', '待收集', '西门子官网下载中心', '样板·待收集', '小型变频器'],
    ['变频器', '三菱', 'FR-E800', '待收集', '待收集', '待收集', '三菱官网下载中心', '样板·待收集', 'E系列'],
    ['变频器', '汇川', 'MD500 / MD380', '待收集', '待收集', '待收集', '汇川官网下载中心', '样板·待收集', '通用变频器'],
    ['变频器', '台达', 'VFD-M / VFD-E', '待收集', '待收集', '待收集', '台达官网下载中心', '样板·待收集', '小型变频器'],
    ['步进电机', '雷赛', '57系列(如57CM22)', '待收集', '待收集', '待收集', '雷赛官网下载中心', '样板·待收集', '步进电机常用'],
    ['接近开关', '欧姆龙', 'E2E / E2B', '待收集', '待收集', '待收集', '欧姆龙官网下载中心', '样板·待收集', '电感式接近开关'],
    ['接近开关', '施耐德', 'XS系列', '待收集', '待收集', '待收集', '施耐德官网下载中心', '样板·待收集', '电感式接近开关'],
    ['光电开关', '欧姆龙', 'E3Z', '待收集', '待收集', '待收集', '欧姆龙官网下载中心', '样板·待收集', '光电传感器'],
    ['传感器', '汇川', '各类光电/接近', '待收集', '待收集', '待收集', '汇川官网下载中心', '样板·待收集', ''],
]
for i, row in enumerate(samples):
    rr = i + 2
    for c, val in enumerate(row, 1):
        cell = ws.cell(rr, c, val)
        cell.font = body_font; cell.alignment = wraptop; cell.border = thin_border
    if i % 2 == 1:
        for c in range(1, len(headers)+1):
            ws.cell(rr, c).fill = alt_fill

widths = [12, 10, 24, 12, 12, 12, 20, 13, 22]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64+c)].width = w
ws.freeze_panes = 'A2'

out = '/home/sandbox/.openclaw/workspace/工控资料库/工控元件索引表_起步版.xlsx'
wb.save(out)
print('saved:', out)
